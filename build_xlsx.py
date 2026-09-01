#!/usr/bin/env python3
"""Wedding & engagement budget workbook. All figures in Rp million (Rp mio)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- palette -------------------------------------------------------------
NAVY="1F3864"; BLUE_HD="2E5496"; LIGHT="D9E1F2"; LIGHTER="EAF0FB"
GREENBG="E2EFDA"; AMBERBG="FFF2CC"; NAVYTX="FFFFFF"

INPUT_BLUE=Font(name="Arial",color="0000FF",size=10)
FORMULA_BK=Font(name="Arial",color="000000",size=10)
LINK_GRN  =Font(name="Arial",color="008000",size=10)
BODY      =Font(name="Arial",size=10)
BODY_B    =Font(name="Arial",size=10,bold=True)
SMALL_IT  =Font(name="Arial",size=8,italic=True,color="606060")
HDR_WHITE =Font(name="Arial",size=11,bold=True,color="FFFFFF")
TITLE_F   =Font(name="Arial",size=16,bold=True,color="1F3864")
SUB_F     =Font(name="Arial",size=10,italic=True,color="606060")

thin=Side(style="thin",color="BFBFBF")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center")
LFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center")
MIO='#,##0.0'; MIO0='#,##0'; PCT='0.0%'
def fill(h): return PatternFill("solid",fgColor=h)

# ---- central reference map (verified once, used everywhere) --------------
A_BAL   ="Assumptions!C5"    # current balance
A_COMB  ="Assumptions!C8"    # combined saving / month
A_ERING ="Assumptions!C9"    # engagement ring
A_WRINGS="Assumptions!C10"   # wedding rings
A_CONT  ="Assumptions!C12"   # contingency %
E_EVENT ="Engagement!C12"    # engagement event total
E_PHASE ="Engagement!C14"    # engagement phase total (event+ring)
W_A     ="'Wedding Budget'!C40"
W_B     ="'Wedding Budget'!D40"

def header(ws,row,cols,text_map,bg=BLUE_HD):
    for c in cols:
        cell=ws.cell(row=row,column=c); cell.fill=fill(bg); cell.font=HDR_WHITE
        cell.alignment=CTR; cell.border=BORDER
    for c,t in text_map.items(): ws.cell(row=row,column=c).value=t

wb=Workbook()

# =========================================================================
# 1. DASHBOARD
# =========================================================================
ws=wb.active; ws.title="Dashboard"; ws.sheet_view.showGridLines=False
for col,w in {"A":2,"B":36,"C":17,"D":34,"E":3}.items(): ws.column_dimensions[col].width=w
ws["B2"]="Wedding & Engagement Plan"; ws["B2"].font=TITLE_F
ws["B3"]="All figures in Rp million. Blue = an input you can change · Black = formula · Green = pulled from another tab."
ws["B3"].font=SUB_F; ws.merge_cells("B3:D3")

ws["B5"]="Where the money stands"; ws["B5"].font=Font(name="Arial",size=12,bold=True,color="1F3864")
kpi=[("Current joint balance (Blugether)","="+A_BAL),
     ("Combined saving / month","="+A_COMB),
     ("Engagement phase cost (event + ring)","="+E_PHASE),
     ("Wedding total — Option A","="+W_A),
     ("Wedding total — Option B","="+W_B)]
r=6
for label,formula in kpi:
    ws.cell(r,2,label).font=BODY; ws.cell(r,2).fill=fill(LIGHTER); ws.cell(r,2).border=BORDER; ws.cell(r,2).alignment=LFT
    c=ws.cell(r,3,formula); c.font=LINK_GRN; c.number_format=MIO; c.alignment=RGT; c.fill=fill(LIGHTER); c.border=BORDER
    r+=1

ws["B13"]="Can we afford it, and when?"; ws["B13"].font=Font(name="Arial",size=12,bold=True,color="1F3864")
header(ws,14,[2,3,4],{2:"Milestone",3:"Funds available",4:"Verdict"})
snap=[("Engagement — Nov 2026","=Cashflow!C7+Cashflow!D7",'=IF(C15>='+E_PHASE+',"Funded ✓","Short")'),
      ("Wedding Option A — Aug 2027","=Cashflow!F16",'=IF(C16>='+W_A+',"Funded ✓, surplus "&TEXT(C16-'+W_A+',"#,##0.0"),"Short by "&TEXT('+W_A+'-C16,"#,##0.0"))'),
      ("Wedding Option A — Nov 2027","=Cashflow!F19",'=IF(C17>='+W_A+',"Funded ✓, surplus "&TEXT(C17-'+W_A+',"#,##0.0"),"Short by "&TEXT('+W_A+'-C17,"#,##0.0"))'),
      ("Wedding Option B — Dec 2027","=Cashflow!F20",'=IF(C18>='+W_B+',"Funded ✓","Short by "&TEXT('+W_B+'-C18,"#,##0.0")&" mio")')]
rr=15
for label,f_funds,f_verd in snap:
    ws.cell(rr,2,label).font=BODY; ws.cell(rr,2).border=BORDER; ws.cell(rr,2).alignment=LFT
    c=ws.cell(rr,3,f_funds); c.font=LINK_GRN; c.number_format=MIO; c.border=BORDER; c.alignment=RGT
    v=ws.cell(rr,4,f_verd); v.font=BODY_B; v.border=BORDER; v.alignment=CTR
    rr+=1
ws["B20"]="Read the tabs left → right: Assumptions · Engagement · Wedding Budget (A vs B) · Cashflow · Timeline."
ws["B20"].font=SUB_F; ws.merge_cells("B20:D20")

# =========================================================================
# 2. ASSUMPTIONS  (values live in column C)
# =========================================================================
wa=wb.create_sheet("Assumptions"); wa.sheet_view.showGridLines=False
for col,w in {"A":2,"B":42,"C":16,"D":44}.items(): wa.column_dimensions[col].width=w
wa["B2"]="Assumptions & Levers"; wa["B2"].font=TITLE_F
wa["B3"]="Change any blue cell — every other tab recalculates."; wa["B3"].font=SUB_F
header(wa,4,[2,3,4],{2:"Item",3:"Value",4:"Note"})
rows=[("Current joint balance",50,MIO,"Blugether joint savings today (Sep 2026)"),
      ("Girl — contribution / month",2,MIO,"Monthly transfer into Blugether"),
      ("Boy — contribution / month",3,MIO,"Monthly transfer into Blugether"),
      ("Combined saving / month","=C6+C7",MIO,"Auto: girl + boy"),
      ("Engagement ring (for the girl)",12.8,MIO,"One-off, paid around engagement"),
      ("Wedding rings (both)",9,MIO,"One-off, inside the wedding budget"),
      ("Groom loafers — already bought",1.6,MIO,"Sunk cost — already paid, not from the pool"),
      ("Contingency / buffer",0.10,PCT,"% added on top of the wedding line items")]
r=5
for label,val,fmt,note in rows:
    wa.cell(r,2,label).font=BODY; wa.cell(r,2).border=BORDER; wa.cell(r,2).alignment=LFT
    c=wa.cell(r,3,val); c.border=BORDER; c.alignment=RGT; c.number_format=fmt
    if isinstance(val,str) and val.startswith("="): c.font=FORMULA_BK
    else: c.font=INPUT_BLUE; c.fill=fill(AMBERBG)
    wa.cell(r,4,note).font=SMALL_IT; wa.cell(r,4).border=BORDER; wa.cell(r,4).alignment=LFT
    r+=1
wa["B15"]="Timeline anchors"; wa["B15"].font=Font(name="Arial",size=11,bold=True,color="1F3864")
header(wa,16,[2,3,4],{2:"Anchor",3:"When",4:"Detail"})
anchors=[("Plan start","Sep 2026","Today"),
         ("Engagement","Nov 2026","At the girl's home, ~20 people"),
         ("Wedding window","2027","Subtle wedding — see Timeline tab for dates")]
r=17
for a,b,cc in anchors:
    wa.cell(r,2,a).font=BODY; wa.cell(r,2).border=BORDER
    wa.cell(r,3,b).font=INPUT_BLUE; wa.cell(r,3).border=BORDER; wa.cell(r,3).alignment=CTR; wa.cell(r,3).fill=fill(AMBERBG)
    wa.cell(r,4,cc).font=SMALL_IT; wa.cell(r,4).border=BORDER; wa.cell(r,4).alignment=LFT
    r+=1

# =========================================================================
# 3. ENGAGEMENT
# =========================================================================
we=wb.create_sheet("Engagement"); we.sheet_view.showGridLines=False
for col,w in {"A":2,"B":32,"C":16,"D":48}.items(): we.column_dimensions[col].width=w
we["B2"]="Engagement — Nov 2026"; we["B2"].font=TITLE_F
we["B3"]="At the girl's home · home-cooked meal · ~20 guests · target Rp 5.0 mio (the ring is tracked separately)."
we["B3"].font=SUB_F; we.merge_cells("B3:D3")
header(we,5,[2,3,4],{2:"Item",3:"Cost",4:"Detail"})
eng=[("Tenda + chairs (rental)",1.5,"Small tent, 20 chairs, tables — 1 day"),
     ("Decoration & flowers",1.5,"Backdrop, drapes, fresh flowers, lighting"),
     ("Food top-up & drinks",1.0,"Ingredients for the home-cooked spread, snacks, ice, water"),
     ("Seserahan / hantaran trays",0.5,"Trays + wrapping for the exchange gifts"),
     ("Documentation",0.3,"Half-day photographer, or a skilled friend"),
     ("Buffer / misc",0.2,"Small unexpected extras")]
r=6
for label,val,note in eng:
    we.cell(r,2,label).font=BODY; we.cell(r,2).border=BORDER; we.cell(r,2).alignment=LFT
    c=we.cell(r,3,val); c.font=INPUT_BLUE; c.fill=fill(AMBERBG); c.number_format=MIO; c.border=BORDER; c.alignment=RGT
    we.cell(r,4,note).font=SMALL_IT; we.cell(r,4).border=BORDER; we.cell(r,4).alignment=LFT
    r+=1
we.cell(12,2,"Engagement event total").font=BODY_B; we.cell(12,2).fill=fill(LIGHT); we.cell(12,2).border=BORDER
t=we.cell(12,3,"=SUM(C6:C11)"); t.font=BODY_B; t.fill=fill(LIGHT); t.number_format=MIO; t.border=BORDER; t.alignment=RGT
we.cell(12,4,"Target was Rp 5.0 mio").font=SMALL_IT; we.cell(12,4).fill=fill(LIGHT); we.cell(12,4).border=BORDER
we.cell(13,2,"vs target (5.0)").font=BODY; we.cell(13,2).border=BORDER
d=we.cell(13,3,"=5-C12"); d.font=FORMULA_BK; d.number_format=MIO; d.border=BORDER; d.alignment=RGT
we.cell(13,4,"Positive = under target").font=SMALL_IT; we.cell(13,4).border=BORDER
we.cell(14,2,"Engagement PHASE total (event + ring)").font=BODY_B; we.cell(14,2).fill=fill(GREENBG); we.cell(14,2).border=BORDER
p=we.cell(14,3,"=C12+"+A_ERING); p.font=BODY_B; p.fill=fill(GREENBG); p.number_format=MIO; p.border=BORDER; p.alignment=RGT
we.cell(14,4,'="Ring Rp "&TEXT('+A_ERING+',"#,##0.0")&" mio added"').font=SMALL_IT
we.cell(14,4).fill=fill(GREENBG); we.cell(14,4).border=BORDER

# =========================================================================
# 4. WEDDING BUDGET  (Option A vs Option B)
# =========================================================================
ww=wb.create_sheet("Wedding Budget"); ww.sheet_view.showGridLines=False
for col,w in {"A":2,"B":30,"C":15,"D":15,"E":42}.items(): ww.column_dimensions[col].width=w
ww["B2"]="Wedding Budget — Option A vs Option B"; ww["B2"].font=TITLE_F
ww["B3"]="Subtle Javanese-adat wedding at Rooang. Pre-wed kept to the siraman; full reception day (MC, music, hotel). A = intimate/lean · B = grander. Rp mio."
ww["B3"].font=SUB_F; ww.merge_cells("B3:E3")
header(ww,5,[2,3,4,5],{2:"Key drivers",3:"Option A",4:"Option B",5:"Note"},bg=NAVY)
drivers=[("Guest count (reception)",50,100,MIO0,"Subtle = small. Drives catering & souvenirs"),
         ("Catering rate / guest",0.15,0.20,MIO,"Per head — Dharmawangsa-area boutique"),
         ("Souvenir rate / guest",0.015,0.025,MIO,"Per favour")]
r=6
for label,a,b,fmt,note in drivers:
    ww.cell(r,2,label).font=BODY; ww.cell(r,2).border=BORDER; ww.cell(r,2).alignment=LFT
    for col,val in ((3,a),(4,b)):
        c=ww.cell(r,col,val); c.font=INPUT_BLUE; c.fill=fill(AMBERBG); c.border=BORDER; c.alignment=RGT; c.number_format=fmt
    ww.cell(r,5,note).font=SMALL_IT; ww.cell(r,5).border=BORDER; ww.cell(r,5).alignment=LFT
    r+=1
header(ww,10,[2,3,4,5],{2:"Line item",3:"Option A",4:"Option B",5:"What it covers"})
items=[("Siraman ceremony","3","5","Setaman flowers, seven-spring water, kendi, dawet, tumpeng, simple gendhing"),
       ("Siraman family catering","1.5","3","Home meal for family on the siraman day"),
       ("Venue — Rooang Dharmawangsa","15","25","A: reception (akad at home) · B: full akad + reception"),
       ("Panggih (temu manten) props","2","3.5","Kembar mayang ×2, suruh, egg, kacar-kucur, sindur"),
       ("Penghulu / KUA (akad)","0.6","0.6","Officiant outside office hours"),
       ("Day-of coordinator (akad + reception)","3","5","Koordinator hari-H — runs the rundown, cues vendors"),
       ("MC / host (akad + reception)","2.5","4","Bilingual host keeping the flow"),
       ("Reception catering","=C6*C7","=D6*D7","Guests × rate / guest"),
       ("Reception decor & floral — Rooang","5","12","Subtle & elegant: stage, entrance, aisle, table florals"),
       ("Entertainment — live music","3.5","8","A: acoustic trio · B: full band"),
       ("Hotel room — bridal suite (1 night)","2.5","5","Wedding-night room for the couple"),
       ("Wedding cake & dessert","1","2.5","Cake + sweet corner"),
       ("Souvenirs / favours","=C6*C8","=D6*D8","Guests × rate / guest"),
       ("Photo & video (2-day)","6","12","Siraman day + wedding day"),
       ("Wedding rings (both)","="+A_WRINGS,"="+A_WRINGS,"Linked from Assumptions"),
       ("Bride — paes MUA + adat busana","10","18","Paes ageng across siraman, akad, reception"),
       ("Groom — adat busana","4","6","Wedding day: beskap, blangkon, keris. Engagement: light grey double-breasted suit. Loafers already bought"),
       ("Family adat attire","2","4","Coordinated kebaya/beskap for the parents"),
       ("Invitations","0.3","1.5","A: digital · B: digital + printed"),
       ("Mahar (dowry)","1","3","Symbolic — set to your intention"),
       ("Transport & misc","2","3.5","Two-day logistics, cars, tips")]
r=11; first=r
for label,a,b,note in items:
    ww.cell(r,2,label).font=BODY; ww.cell(r,2).border=BORDER; ww.cell(r,2).alignment=LFT
    for col,val in ((3,a),(4,b)):
        c=ww.cell(r,col,val); c.border=BORDER; c.alignment=RGT; c.number_format=MIO
        if isinstance(val,str) and val.startswith("="):
            c.font=LINK_GRN if "Assumptions" in val else FORMULA_BK
        else:
            c.font=INPUT_BLUE; c.fill=fill(AMBERBG)
    ww.cell(r,5,note).font=SMALL_IT; ww.cell(r,5).border=BORDER; ww.cell(r,5).alignment=LFT
    r+=1
last=r-1
sub=r
ww.cell(sub,2,"Subtotal").font=BODY_B; ww.cell(sub,2).fill=fill(LIGHT); ww.cell(sub,2).border=BORDER
for col in (3,4):
    L="C" if col==3 else "D"
    c=ww.cell(sub,col,f"=SUM({L}{first}:{L}{last})"); c.font=BODY_B; c.fill=fill(LIGHT); c.number_format=MIO; c.border=BORDER; c.alignment=RGT
ww.cell(sub,5,"Sum of the line items above").font=SMALL_IT; ww.cell(sub,5).fill=fill(LIGHT); ww.cell(sub,5).border=BORDER
con=r+1
ww.cell(con,2,"Contingency").font=BODY; ww.cell(con,2).border=BORDER
for col in (3,4):
    L="C" if col==3 else "D"
    c=ww.cell(con,col,f"={L}{sub}*{A_CONT}"); c.font=FORMULA_BK; c.number_format=MIO; c.border=BORDER; c.alignment=RGT
ww.cell(con,5,'="Buffer @ "&TEXT('+A_CONT+',"0%")').font=SMALL_IT; ww.cell(con,5).border=BORDER
gt=40
ww.cell(gt,2,"WEDDING GRAND TOTAL").font=Font(name="Arial",size=11,bold=True,color="FFFFFF"); ww.cell(gt,2).fill=fill(NAVY); ww.cell(gt,2).border=BORDER
for col in (3,4):
    L="C" if col==3 else "D"
    c=ww.cell(gt,col,f"={L}{sub}+{L}{con}"); c.font=Font(name="Arial",size=11,bold=True,color="FFFFFF"); c.fill=fill(NAVY); c.number_format=MIO; c.border=BORDER; c.alignment=RGT
ww.cell(gt,5,"Subtotal + contingency — the number to fund").font=Font(name="Arial",size=8,italic=True,color="FFFFFF"); ww.cell(gt,5).fill=fill(NAVY); ww.cell(gt,5).border=BORDER
ww.cell(gt+1,2,"Option B minus Option A").font=BODY
db=ww.cell(gt+1,3,f"=D{gt}-C{gt}"); db.font=FORMULA_BK; db.number_format=MIO; db.alignment=RGT
ww.cell(gt+1,4,"extra for B").font=SMALL_IT

# =========================================================================
# 5. CASHFLOW
# =========================================================================
wc=wb.create_sheet("Cashflow"); wc.sheet_view.showGridLines=False
for col,w in {"A":2,"B":13,"C":15,"D":13,"E":16,"F":15,"G":26}.items(): wc.column_dimensions[col].width=w
wc["B2"]="Cashflow projection"; wc["B2"].font=TITLE_F
wc["B3"]="Opening + saving − planned outflow = closing. The engagement outflow lands in Nov 2026. All Rp mio."
wc["B3"].font=SUB_F; wc.merge_cells("B3:G3")
header(wc,4,[2,3,4,5,6,7],{2:"Month",3:"Opening",4:"Saving",5:"Engagement out",6:"Closing",7:"Milestone"})
months=[("Sep 2026",""),("Oct 2026",""),("Nov 2026","Engagement + ring"),("Dec 2026",""),
        ("Jan 2027",""),("Feb 2027",""),("Mar 2027",""),("Apr 2027",""),
        ("May 2027",""),("Jun 2027","← Option A window"),("Jul 2027",""),("Aug 2027",""),
        ("Sep 2027","← Option A window"),("Oct 2027",""),("Nov 2027",""),("Dec 2027","← Option B window")]
r=5
for i,(m,mile) in enumerate(months):
    wc.cell(r,2,m).font=BODY; wc.cell(r,2).border=BORDER; wc.cell(r,2).alignment=CTR
    op="="+A_BAL if i==0 else f"=F{r-1}"
    o=wc.cell(r,3,op); o.font=LINK_GRN if i==0 else FORMULA_BK; o.number_format=MIO; o.border=BORDER; o.alignment=RGT
    s=wc.cell(r,4,"="+A_COMB); s.font=LINK_GRN; s.number_format=MIO; s.border=BORDER; s.alignment=RGT
    if i==2:
        e=wc.cell(r,5,"="+E_PHASE); e.font=LINK_GRN
    else:
        e=wc.cell(r,5,0); e.font=FORMULA_BK
    e.number_format=MIO; e.border=BORDER; e.alignment=RGT
    cl=wc.cell(r,6,f"=C{r}+D{r}-E{r}"); cl.font=BODY_B; cl.number_format=MIO; cl.border=BORDER; cl.alignment=RGT; cl.fill=fill(GREENBG)
    wc.cell(r,7,mile).font=SMALL_IT; wc.cell(r,7).border=BORDER; wc.cell(r,7).alignment=LFT
    r+=1
wc.cell(r+1,2,"Note").font=BODY_B
wc.cell(r+1,3,"Closing balance = money you could put on the wedding that month. Compare it to the Wedding Budget grand totals.").font=SMALL_IT
wc.merge_cells(start_row=r+1,start_column=3,end_row=r+1,end_column=7)

# =========================================================================
# 6. TIMELINE
# =========================================================================
wt=wb.create_sheet("Timeline"); wt.sheet_view.showGridLines=False
for col,w in {"A":2,"B":22,"C":13,"D":15,"E":15,"F":15,"G":34}.items(): wt.column_dimensions[col].width=w
wt["B2"]="Timeline scenarios"; wt["B2"].font=TITLE_F
wt["B3"]="Same savings plan, three wedding dates. 'Funds available' is that month's closing balance from Cashflow."
wt["B3"].font=SUB_F; wt.merge_cells("B3:G3")
header(wt,5,[2,3,4,5,6,7],{2:"Scenario",3:"Wedding",4:"Funds avail.",5:"Option A gap",6:"Option B gap",7:"Read"})
scen=[("1 · Intimate & earliest","Aug 2027","=Cashflow!F16"),
      ("2 · Comfortable cushion","Nov 2027","=Cashflow!F19"),
      ("3 · Full grandeur (Opt B)","Dec 2027","=Cashflow!F20")]
r=6
for name,when,funds in scen:
    wt.cell(r,2,name).font=BODY_B; wt.cell(r,2).border=BORDER; wt.cell(r,2).alignment=LFT
    wt.cell(r,3,when).font=BODY; wt.cell(r,3).border=BORDER; wt.cell(r,3).alignment=CTR
    fc=wt.cell(r,4,funds); fc.font=LINK_GRN; fc.number_format=MIO; fc.border=BORDER; fc.alignment=RGT
    ga=wt.cell(r,5,f"=D{r}-{W_A}"); gb=wt.cell(r,6,f"=D{r}-{W_B}")
    for c in (ga,gb): c.number_format=MIO; c.border=BORDER; c.alignment=RGT; c.font=FORMULA_BK
    read=wt.cell(r,7,f'=IF(E{r}>=0,IF(F{r}>=0,"Both A & B funded ✓","Option A funded; B short"),"A short — add gifts/angpau or more months")')
    read.font=SMALL_IT; read.border=BORDER; read.alignment=LFT
    r+=1
wt.cell(11,2,"Gap = funds available − wedding grand total.  Positive = surplus, negative = shortfall.").font=SMALL_IT; wt.merge_cells("B11:G11")
wt.cell(12,2,"Upside not modelled: angpau/gift money at the reception often offsets 30–50% of catering. Treat it as a cushion, not a plan.").font=SMALL_IT; wt.merge_cells("B12:G12")
wt.cell(13,2,"Recommendation").font=BODY_B
wt.cell(14,2,"Option A — siraman + a full reception day at an intimate scale — clears around August 2027 and sits on a cushion by Oct–Nov. Option B (grander, ~100 guests) is a 2028 project, or needs a bigger monthly saving plus angpau.").font=BODY
wt.cell(14,2).alignment=LFT; wt.merge_cells("B14:G14"); wt.row_dimensions[14].height=32

out="/Users/gomobile/Documents/Project/Wedding-Budget/Wedding_Budget.xlsx"
wb.save(out); print("saved",out)
